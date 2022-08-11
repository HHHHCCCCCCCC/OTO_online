import math
import serial
from ctypes import *
import openpyxl
import datetime
import pandas as pd
import numpy as np
from pre_spectrum import pretreatment as pre



class IRConfig():
    def __init__(self):
        self.ComFormat = 'M{:0>6s}{:1s}{:0>2s}{:0>2s}'
        self.ComParity = {'None': '0', 'Odd': '1', 'Even': '2'}
        self.wb = openpyxl.load_workbook('config.xlsx')
        self.Work_Excel = self.wb.active
        self.Work_Excel = self.wb['光谱仪设置']
        self.comnumber = 'COM' + str(self.Work_Excel.cell(2, 2).value)
        # print(self.comnumber)
        self.ModbusConfig = self.ComFormat.format(str(self.Work_Excel.cell(3, 2).value),
                                                  self.ComParity[str(self.Work_Excel.cell(5, 2).value)],
                                                  str(self.Work_Excel.cell(8, 2).value), str(self.Work_Excel.cell(9, 2).value))
        self.ScanTimes = (self.Work_Excel.cell(13, 2).value)
        self.IntergrationTime = int(self.Work_Excel.cell(14, 2).value) # 积分时间 （us）
        print(self.IntergrationTime)
        self.LoopTime = self.Work_Excel.cell(15, 2).value  # 间隔时间
        self.TemperatureSet = self.Work_Excel.cell(16, 2).value
        if str(self.Work_Excel.cell(18, 2).value) == "ON":
            self.TECONOFF = 1
        else:
            self.TECONOFF = 0
        # print(self.TemperatureSet)
        self.Pre_Function = [''] * 10
        self.beta = [0] * 216
        self.background = [0] * 216
        self.MA_WinSize = 0
        self.SG_WinLen = 0
        self.SG_PolyOrder = 0
        self.spline_xmin = 0
        self.spline_numbers = 0
        self.spline_xmax = 0
        Pre_Cnt = 0
        while True:
            if str(self.Work_Excel.cell(21 + Pre_Cnt, 1).value) != 'end':
                self.Pre_Function[Pre_Cnt] = str(self.Work_Excel.cell(21 + Pre_Cnt, 1).value)
                if self.Pre_Function[Pre_Cnt] == 'smoothMA':
                    self.MA_WinSize = int(self.Work_Excel.cell(21 + Pre_Cnt, 2).value)
                elif self.Pre_Function[Pre_Cnt] == 'Diff':
                    self.SG_WinLen = int(self.Work_Excel.cell(21 + Pre_Cnt, 2).value)
                    self.SG_PolyOrder = int(self.Work_Excel.cell(21 + Pre_Cnt, 3).value)
                elif self.Pre_Function[Pre_Cnt] == 'Spline':
                    self.spline_xmin = int(self.Work_Excel.cell(21 + Pre_Cnt, 2).value)
                    self.spline_xmax = int(self.Work_Excel.cell(21 + Pre_Cnt, 3).value)
                    self.spline_numbers = int(self.Work_Excel.cell(21 + Pre_Cnt, 4).value)
                Pre_Cnt = Pre_Cnt + 1
            else:
                break
        # print(self.MA_WinSize )

        self.Head_Information = [
            ['#######%Modbus协议设置###############'],
            ['1.串口号：', str(self.Work_Excel.cell(2, 2).value)],
            ['2.波特率：', str(self.Work_Excel.cell(3, 2).value)],
            ['3.数据位：',  str(self.Work_Excel.cell(4, 2).value)],
            ['4.校验位：',  str(self.Work_Excel.cell(5, 2).value)],
            ['5.停止位：',  str(self.Work_Excel.cell(6, 2).value)],
            ['6.超时时间',  str(self.Work_Excel.cell(7, 2).value)],
            ['7.设备号：',  str(self.Work_Excel.cell(8, 2).value)],
            ['8.寄存器开始：', str(self.Work_Excel.cell(9, 2).value)],
            ['9.寄存器个数：', str(self.Work_Excel.cell(10, 2).value)],
            [' '],
            ['#########%光谱仪设置：第一个是扫描次数，每次都是n次平均后的结果并保留#########'],
            ['1.扫描次数：', str(self.Work_Excel.cell(13, 2).value)],
            ['2. 积分时间（us)', str(self.Work_Excel.cell(14, 2).value)],
            ['3.间隔时间（s)：', str(self.Work_Excel.cell(15, 2).value)],
            ['4.探测器温度设置（℃)：:', str(self.Work_Excel.cell(16, 2).value)],
            ['5.光源工作模式：（LC/LO)', str(self.Work_Excel.cell(17, 2).value)],
            ['6.探测器制冷（ON/OFF） ',str(self.Work_Excel.cell(18, 2).value)],
            [' '],
            [' '],
            [' '],
            ['#########%Pretreat      """光谱预处理按下列顺序执行操作:标准化；归一化；求导；差分···以end为标记结束"""#########']
        ]

        self.sheet2 = self.wb['β系数']                                     #读取β系数
        self.beta_constant = self.sheet2.cell(1,2).value
        for i in range(0, 216):  #249个波长,另加一个beta常量
            self.beta[i] = float(self.sheet2.cell(2 + i, 2).value)

        self.sheet3 = self.wb['背景光谱']                                    #读取背景光谱
        self.background_constant = self.sheet3.cell(1,2).value
        for i in range(0, 216):  #249个波长,另加一个beta常量
            self.background[i] = float(self.sheet3.cell(1 + i, 2).value)
        # self.sheet4 = self.wb['温度修正']                                   # 读取温度比例
        self.re_ex = pd.read_excel('config.xlsx', sheet_name='温度修正',header=None)
        self.re_ex = np.array(self.re_ex)
        self.Temperaturecorrection_T = self.re_ex[0,:]
        self.Temperaturecorrection_mutiple = self.re_ex[1:, :]
        self.wb.close()
        # self.beta = np.array(self.beta)
        # self.background = np.array(self.background)
        # sub = self.beta - self.background
        # print(self.background)

class Device:
    def __init__(self):
        self.VID = 1592
        self.PID = 2732
        # self.receive_frame_last = 20
        self.pre = pre()
        self.errStatus = 0
        self.contrl_com_enable = 0
        self.OTOdll = CDLL("UserApplication.dll")
        self.LoadConfig = IRConfig()
        self.intFramesize = self.GetFramesize()
        self.Temperature = self.TEC()
        self.IsCOM_Communication()
        self.backup_cnt = 0

        # print(self.LoadConfig.Head_Information)

    def IsConnected(self):
        #Check how many device is connected with PC.
        self.intDeviceamout = c_int(0)
        self.OTOdll.UAI_SpectrometerGetDeviceAmount(self.VID,self.PID,byref(self.intDeviceamout))
        return self.intDeviceamout.value

    def IsCOM_Connected(self):
        if self.contrl_com_enable == 1:
            return 1
        else:
            return 0

    def IsCOM_Communication(self):
        if self.contrl_com_enable == 1:
            count = self.contrl_com.inWaiting()                #获取串口缓冲区数据
            self.contrl_com.flushInput()  # 清空缓冲区
            if count != 0:
                self.receive_frame = self.contrl_com.readline(count)
                # print(self.receive_frame)
                # self.contrl_com.flushInput()
        else:
            try:
                self.COM_Control = (self.LoadConfig.comnumber)
                self.contrl_com = serial.Serial(self.COM_Control , 9600, timeout=1)
                self.receive_frame = self.contrl_com.readline()
                self.contrl_com_enable = 1
            except:
                self.contrl_com_enable = 0
                return 0
        return self.receive_frame
                # self.label_COM_connected.setText('串口打开失败!')

    def Contrl_Communication(self, data):
        if(self.contrl_com_enable == 1):
            self.contrl_com.write(data.encode())
            # print(data)
        else:
            self.contrl_com_enable = 0
            # QMessageBox.information(self,  "提示", "数据接收错误！")
            # self.IsCOM_Connected()

    def Open(self):
        #Open Device
        self.DeviceHandle = c_int(0)
        self.OTOdll.UAI_SpectrometerOpen(0,byref(self.DeviceHandle),self.VID,self.PID)
        return self.DeviceHandle.value

    def GetFramesize(self):
        self.Open()
        #Get Framesize
        self.intFramesize = c_int(0)
        self.OTOdll.UAI_SpectromoduleGetFrameSize(self.DeviceHandle,byref(self.intFramesize))
        print("Device framesize:")
        return self.intFramesize

    def GetModuleName(self):
        #Get Module name
        self.charModulename = create_string_buffer(16)
        self.OTOdll.UAI_SpectrometerGetModelName(self.DeviceHandle,byref(self.charModulename))
        # print("Module name:")
        return repr(self.charModulename.value)

    def GetSerilaNum(self):
        #Get Serial number
        self.charSerialnumber = create_string_buffer(16)
        self.OTOdll.UAI_SpectrometerGetSerialNumber(self.DeviceHandle,byref(self.charSerialnumber))
        print("Serial number:")
        return repr(self.charSerialnumber.value)

    def TEC(self):
        # self.GetTECOnOff = c_int(0)
        self.OTOdll.UAI_SpectrometerSetTECOnOff(self.DeviceHandle, self.LoadConfig.TECONOFF)
        # self.OTOdll.UAI_SpectrometerGetTECOnOff(self.DeviceHandle, byref(self.GetTECOnOff))
        self.OTOdll.UAI_SpectrometerSetTECTargetTemperature(self.DeviceHandle, c_float(self.LoadConfig.TemperatureSet))
        # self.OTOdll.UAI_SpectrometerSetTECDAC(self.DeviceHandle, 0xff)

    def GetTemperature(self):
        self.temperature = c_float(0)
        self.OTOdll.UAI_SpectrometerGetTECTemperature(self.DeviceHandle, byref(self.temperature))
        # print(self.temperature)
        return repr(self.temperature.value)


    def GetLambda(self):
        self.TempLambda = (c_float*self.intFramesize.value)()
        #Get wavelength
        self.OTOdll.UAI_SpectrometerWavelengthAcquire(self.DeviceHandle,byref(self.TempLambda))
        self.Lambda = []
        for i in range(0, self.intFramesize.value):
            self.Lambda.append(self.TempLambda[i])
        return self.Lambda

    def GetIntensity(self):
        self.TempIntensity = (c_float * self.intFramesize.value)()
        # Get Intensity
        self.OTOdll.UAI_SpectrometerDataOneshots(self.DeviceHandle,self.LoadConfig.IntergrationTime,byref(self.TempIntensity),int(self.LoadConfig.ScanTimes))
        # Do Background
        self.OTOdll.UAI_BackgroundRemove(self.DeviceHandle, self.intFramesize.value, byref(self.TempIntensity))
        # Do Linearity
        self.OTOdll.UAI_LinearityCorrection(self.DeviceHandle, self.intFramesize.value, byref(self.TempIntensity))
        self.Intensity = []
        for i in range(0, self.intFramesize.value):
            self.Intensity.append(self.TempIntensity[i]/self.LoadConfig.IntergrationTime)
        return self.Intensity

    @property
    def Temperature_correction(self):
        excel_filename5 = "D:/Pukon/Pukon_" + str('环境温度') + "/" + '环境温度' + ".xlsx"
        self.Temperature_coef = []
        self.read_last = 1
        try:
            self.receive_frame = self.IsCOM_Communication()                                            #读取温度
            self.receive_frame = float(self.receive_frame)
            print(self.receive_frame)
            self.receive_frame_last = pd.read_excel(excel_filename5, header=None)
            self.receive_frame_last = float(self.receive_frame_last.iloc[-1, 1])
            print(self.receive_frame_last)
            self.difference = self.receive_frame - self.receive_frame_last                           #计算差值
        except:
            self.receive_frame_last = pd.read_excel(excel_filename5,  header=None)
            self.receive_frame = float(self.receive_frame_last.iloc[-1, 1])
            self.read_last = 0
        if math.fabs(self.difference) > 5:
            self.receive_frame = self.receive_frame_last
            self.read_last = 0
        else:
            pass
        # print(self.receive_frame)
        # self.current = str(datetime.datetime.now().replace(microsecond=0))
        # self.current = self.current.replace(':', '-')
        self.Temp_save = self.receive_frame
        self.save_Temperature(excel_filename5, self.Temp_save)
        self.T_current = self.receive_frame
        for i in range(0, 216):                                    #循环获取多项式系数
            self.x = self.LoadConfig.Temperaturecorrection_T  # 自变量
            self.y = self.LoadConfig.Temperaturecorrection_mutiple[i,:]
            self.x = list(self.x.flatten())
            self.y = list(self.y.flatten())
            self.coef = np.polyfit(self.x, self.y, 3)  # 按三次多项式拟合
            self.Temperature_coef.append(self.coef)
        self.Temperature_coef= np.array(self.Temperature_coef)
        self.Temperature_coef = self.Temperature_coef[:, [3, 2, 1, 0]]
        self.T_ = [[self.T_current ** 0, self.T_current ** 1, self.T_current ** 2, self.T_current ** 3]]
        self.T_ = np.array(self.T_).T
        self.tempera_mutiple = np.dot(self.Temperature_coef, self.T_)  # 线性代数中的矩阵相乘
        self.tempera_mutiple= np.array(self.tempera_mutiple)
        self.tempera_mutiple = list(self.tempera_mutiple.flatten())
        return self.tempera_mutiple


    def SaveExcel(self, excel_name, indesity_data):
        wb = openpyxl.Workbook()
        Work_Excel = wb.active
        for j in range(0, len(self.LoadConfig.Head_Information)):
            Work_Excel.append(self.LoadConfig.Head_Information[j])  # 把每一行append到worksheet中
        for j in range(0, len(self.LoadConfig.Pre_Function)):
            Work_Excel.append([self.LoadConfig.Pre_Function[j]])  # 把每一行append到worksheet中
        for j in range(0, len(self.WaveLenth)):
            Work_Excel.append([self.WaveLenth[j]] + [indesity_data[j]])
        wb.save(excel_name)

    def save_water(self, excel_name, water_data):
        try:
            wb = openpyxl.load_workbook("D:/Pukon/Pukon_" + str('理化指标') + "/" + '理化指标' + ".xlsx")
            Work_Excel = wb.active
            Work_Excel.append(water_data)
            wb.save(excel_name)
        except FileNotFoundError:
            wb_creat = openpyxl.Workbook()
            Work_Excel = wb_creat.active
            Work_Excel.append(water_data)
            wb_creat.save(excel_name)

    def Button_back(self):                                     #保存背景光谱到config文件
        self.Intensitie = self.GetIntensity()
        self.Intensitie_Back = self.Intensitie[16:232]
        self.WaveLenth = self.GetLambda()
        self.WaveLenth = self.WaveLenth[16:232]
        wb = openpyxl.load_workbook("config.xlsx")
        sheet3 = wb['背景光谱']  # 读取背景光谱
        print(1)
        for i in range(0, 216):  # 249个波长,另加一个beta常量
            (sheet3.cell(1 + i, 1).value) = self.WaveLenth[i]
            (sheet3.cell(1 + i, 2).value) = self.Intensitie_Back[i]
        print(sheet3.cell(2, 2).value)
        print(2)
        wb.save("config.xlsx")


    def save_backup(self, excel_name,  excel_name_backup):            #理化指标备份，一百个数据备份一次，防止掉电中断数据丢失
        wb = openpyxl.load_workbook(excel_name)
        sTargetFile = excel_name_backup
        wb.save(sTargetFile)
        self.backup_cnt = 0
        # print("It is over")

    def save_Temperature(self, excel_name, Temperature):
        try:
            wb = openpyxl.load_workbook("D:/Pukon/Pukon_" + str('环境温度') + "/" + '环境温度' + ".xlsx")
            Work_Excel = wb.active
            Work_Excel.cell(row=1, column=2, value=Temperature)
            # Work_Excel(Temperature)
            wb.save(excel_name)
        except FileNotFoundError:
            wb_creat = openpyxl.Workbook()
            Work_Excel = wb_creat.active
            Work_Excel.append(Temperature)
            wb_creat.save(excel_name)

    def WriteDataToExcl(self):
        self.current = str(datetime.datetime.now().replace(microsecond=0))
        self.current = self.current.replace(':', '-')
        self.Intensitie = self.GetIntensity()
        self.Intensitie = self.Intensitie[16:232]
        self.WaveLenth = self.GetLambda()
        self.WaveLenth = self.WaveLenth[16:232]

        # 保存原始光谱
        excel_filename0 = "D:/Pukon/Pukon_" + str('原始光谱') + "/" + self.current + '-原始光谱' + ".xlsx"
        self.SaveExcel(excel_filename0,  self.Intensitie)
        self.temperature_correction = self.Temperature_correction

        # 保存环境温度
        self.Intensities = []
        # self.Intensi = []
        for i in range(len(self.Intensitie)):
            self.Intensi = self.Intensitie[i] * self.temperature_correction[i]
            self.Intensities.append(self.Intensi)
        # print(self.Intensities)
        self.Absorb = 216 * [0]
        # print(self.WaveLenth)
        self.Temperature = self.GetTemperature()
        self.LoadConfig.Head_Information[18] = ['7.当前制冷温度（℃）', self.Temperature]

        #文件名
        excel_filename1 = "D:/Pukon/Pukon_" + str('透射光谱') + "/" + self.current + '-透射光谱' + ".xlsx"
        excel_filename2 = "D:/Pukon/Pukon_" + str('吸收光谱') + "/" + self.current + '-吸收光谱' + ".xlsx"
        excel_filename3 = "D:/Pukon/Pukon_" + str('吸收预处理') + "/" + self.current + '-吸收预处理' + ".xlsx"
        excel_filename4 = "D:/Pukon/Pukon_" + str('理化指标') + "/" + '理化指标' + ".xlsx"
        excel_filename5 = "D:/Pukon/Pukon_" + str('理化指标') + "/" + '理化指标备份' + ".xlsx"

        # 用于保留原始温度修正光谱
        self.Intensities_transmission = np.array(self.Intensities.copy())
        self.SaveExcel(excel_filename1,  self.Intensities_transmission)

        # 用于保留吸收光谱
        # print(self.LoadConfig.background)
        # print(self.Intensities)
        for i in range(0, len(self.Intensities)):
            self.Absorb[i] = math.log10(self.LoadConfig.background[i]/self.Intensities[i])
        # print(self.Absorb)
        self.SaveExcel(excel_filename2, self.Absorb)

        # 用于保留吸收预处理光谱
        # 第一个预处理方法代入吸收光谱数据
        for i in range(0, len(self.LoadConfig.Pre_Function)):
            if self.LoadConfig.Pre_Function[i] != '':
                # excel_filename = "D:/Pukon/Pukon_" + str(self.LoadConfig.Pre_Function[i]) + "/" + self.current + '-' \
                #                  + str(self.LoadConfig.Pre_Function[i]) + ".xlsx"
                print(self.LoadConfig.Pre_Function[i])
                # print(i)
                if i==0:
                    self.Intensities_Temp = self.Absorb
                else:
                    self.Intensities_Temp = self.Intensities_Temp
                if str(self.LoadConfig.Pre_Function[i]) == 'SNV':
                    self.Intensities_Temp = self.pre.SNV(self.Intensities_Temp)
                elif str(self.LoadConfig.Pre_Function[i]) == 'Diff':
                    self.Intensities_Temp = self.pre.Diff(self.Intensities_Temp, self.LoadConfig.SG_WinLen,self.LoadConfig.SG_PolyOrder)

                elif str(self.LoadConfig.Pre_Function[i]) == 'smoothMA':
                    self.Intensities_Temp = self.pre.smoothMA(self.Intensities_Temp, self.LoadConfig.MA_WinSize)

                elif str(self.LoadConfig.Pre_Function[i]) == 'MaxNorm':
                    self.Intensities_Temp = self.pre.MaxNorm(self.Intensities_Temp)

                elif str(self.LoadConfig.Pre_Function[i]) == 'MaxMinNormalize':
                    self.Intensities_Temp = self.pre.MaxMinNormalize(self.Intensities_Temp)

                elif str(self.LoadConfig.Pre_Function[i]) == 'MeanCenter':
                    self.Intensities_Temp = self.pre.MeanCenter(self.Intensities_Temp)
                elif str(self.LoadConfig.Pre_Function[i]) == 'vector_Normalize':
                    self.Intensities_Temp = self.pre.vector_Normalize(self.Intensities_Temp)

                # if str(self.LoadConfig.Pre_Function[i]) == 'Spline':
                #     print(123)
                #     self.WaveLenth ,self.Intensities = self.pre.Spline( self.WaveLenth,self.Intensities ,self.LoadConfig.spline_xmin ,self.LoadConfig.spline_xmax , self.LoadConfig.spline_numbers)
        self.Intensities_Temp = np.array(self.Intensities_Temp).flatten()
        self.Intensities_Temp = self.Intensities_Temp.tolist()
        self.SaveExcel(excel_filename3, self.Intensities_Temp)

        # 用于保存理化指标
        self.water = self.LoadConfig.beta_constant
        for i in range(0, len(self.Intensities)):
            self.water += self.Intensities_Temp[i] * self.LoadConfig.beta[i]
        self.water = [self.current, self.water]
        self.save_water(excel_filename4, self.water)
        # print(self.water)
        self_WaterData = '%.2f' % (int(self.water[1] * 100) * 0.01)
        if len(self_WaterData) != 5:
            self_WaterData = self_WaterData.zfill(5)
        string_temp = 'W' + str(self_WaterData ) + '\r' + '\n'
        # self_WaterData = '%.2f' % (int(self.water[1] * 100) * 0.01)
        if 0<float(self_WaterData)<=100:
            self.Contrl_Communication(string_temp)
        else:
            self.Contrl_Communication('NaN\r\n')
        self.backup_cnt += 1
        if self.backup_cnt == 100:
            self.save_backup(excel_filename4,excel_filename5)      #理化指标备份，一百个数据备份一次，防止掉电中断数据丢失
        return self.water


